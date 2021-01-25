import json

from collections.abc import MutableMapping, Iterable
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.cell import WriteOnlyCell
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.renderers import BaseRenderer
from rest_framework.utils.serializer_helpers import ReturnDict, ReturnList
from django.core.serializers.json import DjangoJSONEncoder


def get_style_from_dict(style_dict, style_name):
    """
    Make NamedStyle instance from dictionary
    :param style_dict: dictionary with style properties.
           Example:    {'fill': {'fill_type'='solid',
                                 'start_color'='FFCCFFCC'},
                        'alignment': {'horizontal': 'center',
                                      'vertical': 'center',
                                      'wrapText': True,
                                      'shrink_to_fit': True},
                        'border_side': {'border_style': 'thin',
                                        'color': 'FF000000'},
                        'font': {'name': 'Arial',
                                 'size': 14,
                                 'bold': True,
                                 'color': 'FF000000'}
                        }
    :param style_name: name of created style
    :return: openpyxl.styles.NamedStyle instance
    """
    style = NamedStyle(name=style_name)
    if not style_dict:
        return style
    for key, value in style_dict.items():
        if key == "font":
            style.font = Font(**value)
        elif key == "fill":
            style.fill = PatternFill(**value)
        elif key == "alignment":
            style.alignment = Alignment(**value)
        elif key == "border_side":
            side = Side(**value)
            style.border = Border(left=side, right=side, top=side, bottom=side)

    return style


def get_attribute(get_from, prop_name, default=None):
    """
    Get attribute from object with name <prop_name>, or take it from function get_<prop_name>
    :param get_from: instance of object
    :param prop_name: name of attribute (str)
    :param default: what to return if attribute doesn't exist
    :return: value of attribute <prop_name> or default
    """
    prop = getattr(get_from, prop_name, None)
    if not prop:
        prop_func = getattr(get_from, "get_{}".format(prop_name), None)
        if prop_func:
            prop = prop_func()
    if prop is None:
        prop = default
    return prop


class XLSXRenderer(BaseRenderer):
    """
    Renderer for Excel spreadsheet open data format (xlsx).
    """

    media_type = "application/xlsx"
    format = "xlsx"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Render `data` into XLSX workbook, returning a workbook.
        """
        if not self._check_validatation_data(data):
            return self._json_format_response(data)

        if data is None:
            return bytes()

        wb = Workbook(write_only=True)
        self.ws = wb.create_sheet()

        results = data["results"] if "results" in data else data

        # Retrieve the configuration parameters from the view
        header = get_attribute(renderer_context["view"], "xlsx_header", {})
        column_header = get_attribute(renderer_context["view"], "xlsx_column_header", {})
        self.body = get_attribute(renderer_context["view"], "xlsx_body", {})

        header_style = get_style_from_dict(header.get("style"), "header_style")
        column_header_style = get_style_from_dict(column_header.get("style"),
                                                  "column_header_style")

        self.body_style = get_style_from_dict(self.body.get("style"), "body_style")

        img_addr = header.get("img")
        if img_addr:
            img = Image(img_addr)
            self.ws.add_image(img, "A1")

        self.ws.title = header.get("tab_title", "Report")
        header_title = header.get("header_title", "Report")

        column_count = 0

        # Set the header row
        if header:
            cell = WriteOnlyCell(self.ws, value=header_title)
            cell.style = header_style
            self.ws.append([cell])


        # Make column headers
        column_titles = column_header.get("titles", [])

        # If we have results, pull the columns names from the keys of the first row
        if len(results):
            if isinstance(results, (ReturnDict, dict)):
                column_names_first_row = results
            elif isinstance(results, (ReturnList, list)):
                column_names_first_row = self._flatten(results[0])

            column_header_row = []
            for column_name in column_names_first_row.keys():
                if column_name == "row_color":
                    continue
                column_count += 1
                if column_count > len(column_titles):
                    column_name_display = column_name
                else:
                    column_name_display = column_titles[column_count - 1]

                cell = WriteOnlyCell(self.ws, value=column_name_display)
                cell.style = column_header_style

                column_header_row.append(cell)

            self.ws.append(column_header_row)

        # Make body

        if isinstance(results, ReturnDict):
            self._make_body(results)
        elif isinstance(results, (ReturnList, list)):
            for row in results:
                self._make_body(row)

        return save_virtual_workbook(wb)


    def _check_validatation_data(self, data):
        detail_key = "detail"
        if detail_key in data:
            return False
        return True


    def _flatten(self, data, parent_key="", key_sep=".", list_sep=", "):
        items = []
        for k, v in data.items():
            new_key = f"{parent_key}{key_sep}{k}" if parent_key else k
            if isinstance(v, MutableMapping):
                items.extend(self._flatten(v, new_key, key_sep=key_sep).items())
            elif isinstance(v, Iterable) and not isinstance(v, str):
                if len(v) > 0 and isinstance(v[0], Iterable):
                    # array of array; write as json
                    items.append((new_key, json.dumps(v, cls=DjangoJSONEncoder)))
                else:
                    # Flatten the array into a comma separated string to fit
                    # in a single spreadsheet column
                    items.append((new_key, list_sep.join(map(str, v))))
            else:
                items.append((new_key, v))
        return dict(items)


    def _json_format_response(self, response_data):
        return json.dumps(response_data)


    def _make_body(self, row):

        flatten_row = self._flatten(row)

        if "row_color" in row:
            fill = PatternFill(fill_type="solid", start_color=row["row_color"])
        else:
            fill = None

        data_row = []
        for column_name, value in flatten_row.items():
            if column_name == "row_color":
                continue

            cell = WriteOnlyCell(self.ws, value=value)
            cell.style = self.body_style
            if fill:
                cell.fill = fill
            data_row.append(cell)

        self.ws.append(data_row)
